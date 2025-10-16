import openpyxl

# Función manual para calcular raíz cuadrada (método babilónico)
def raiz_cuadrada(n):
    if n == 0:
        return 0
    x = n
    y = (x + 1) / 2
    while y < x:
        x = y
        y = (x + n / x) / 2
    return x

# Función manual para calcular potencia
def potencia(base, exponente):
    if exponente == 0:
        return 1
    resultado = 1
    for i in range(int(exponente)):
        resultado *= base
    return resultado

# Función para calcular suma de una lista
def suma(lista):
    total = 0
    for valor in lista:
        total += valor
    return total

# Función para calcular media
def media(lista):
    return suma(lista) / len(lista)

# Función para leer archivo Excel
def leer_excel(archivo):
    wb = openpyxl.load_workbook(archivo)
    hoja = wb.active
    
    X = []
    Y = []
    
    # Detectar si la primera fila es encabezado
    primera_fila = list(hoja.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    fila_inicio = 1
    
    # Si la primera celda no es un número, es encabezado
    try:
        float(primera_fila[0])
        fila_inicio = 1
    except (ValueError, TypeError):
        fila_inicio = 2
    
    # Leer datos desde la fila correspondiente
    for fila in hoja.iter_rows(min_row=fila_inicio, values_only=True):
        # Verificar que haya al menos 2 columnas
        if len(fila) < 2:
            continue
            
        if fila[0] is not None and fila[1] is not None:
            try:
                # Convertir a float, manejando diferentes formatos
                x_val = str(fila[0]).strip()
                y_val = str(fila[1]).strip()
                
                # Manejar formato de hora HH:MM a número decimal (horas)
                if ':' in x_val:
                    partes = x_val.split(':')
                    horas = float(partes[0])
                    minutos = float(partes[1]) if len(partes) > 1 else 0
                    x = horas + (minutos / 60.0)  # Convertir a decimal
                else:
                    x = float(x_val)
                
                # Manejar comas decimales
                y_val = y_val.replace(',', '.')
                y = float(y_val)
                
                X.append(x)
                Y.append(y)
            except (ValueError, TypeError):
                continue
    
    return X, Y

# 1. DESVIACIÓN ESTÁNDAR TOTAL
def calcular_desviacion_estandar(Y):
    n = len(Y)
    mean_Y = media(Y)
    
    # Calcular suma de cuadrados de diferencias
    suma_cuadrados = 0
    for yi in Y:
        diferencia = yi - mean_Y
        suma_cuadrados += diferencia * diferencia
    
    # Desviación estándar muestral (n-1)
    varianza = suma_cuadrados / (n - 1)
    desv_std = raiz_cuadrada(varianza)
    
    return desv_std

# 2. COEFICIENTE DE CORRELACIÓN
def calcular_correlacion(X, Y):
    n = len(X)
    mean_X = media(X)
    mean_Y = media(Y)
    
    # Numerador: Σ(xi - x̄)(yi - ȳ)
    numerador = 0
    for i in range(n):
        numerador += (X[i] - mean_X) * (Y[i] - mean_Y)
    
    # Denominador: √[Σ(xi - x̄)² · Σ(yi - ȳ)²]
    suma_cuadrados_x = 0
    suma_cuadrados_y = 0
    
    for i in range(n):
        suma_cuadrados_x += (X[i] - mean_X) * (X[i] - mean_X)
        suma_cuadrados_y += (Y[i] - mean_Y) * (Y[i] - mean_Y)
    
    denominador = raiz_cuadrada(suma_cuadrados_x * suma_cuadrados_y)
    
    r = numerador / denominador
    
    return r

# 3. CALCULAR REGRESIÓN LINEAL (y = a + bx)
def calcular_regresion(X, Y):
    n = len(X)
    mean_X = media(X)
    mean_Y = media(Y)
    
    # Calcular pendiente (b)
    numerador = 0
    denominador = 0
    
    for i in range(n):
        numerador += (X[i] - mean_X) * (Y[i] - mean_Y)
        denominador += (X[i] - mean_X) * (X[i] - mean_X)
    
    b = numerador / denominador  # Pendiente
    a = mean_Y - b * mean_X      # Intercepto
    
    return a, b

# 4. ERROR ESTÁNDAR DEL ESTIMADO
def calcular_error_estandar(X, Y, a, b):
    n = len(X)
    
    # Calcular valores predichos y residuos
    suma_residuos_cuadrados = 0
    
    for i in range(n):
        y_predicho = a + b * X[i]
        residuo = Y[i] - y_predicho
        suma_residuos_cuadrados += residuo * residuo
    
    # Error estándar del estimado (n-2 para regresión simple)
    see = raiz_cuadrada(suma_residuos_cuadrados / (n - 2))
    
    return see

# 5. COEFICIENTE DE DETERMINACIÓN
def calcular_r_cuadrado(r):
    return r * r


def calcular_estadisticas(archivo_excel):
    print(" ANÁLISIS DE REGRESIÓN")
    print()
    
    
    X, Y = leer_excel(archivo_excel)
    n = len(X)
    
    if n < 2:
        print("Error: Se necesitan al menos 2 datos para calcular estadísticas")
        return
    
    
    mean_X = media(X)
    mean_Y = media(Y)
    
    # Desviación estándar total
    desv_std = calcular_desviacion_estandar(Y)
    
    # Coeficiente de correlación
    r = calcular_correlacion(X, Y)
    
    # Coeficiente de determinación
    r2 = calcular_r_cuadrado(r)
    
    # Regresión lineal
    a, b = calcular_regresion(X, Y)
    
    # Error estándar del estimado
    see = calcular_error_estandar(X, Y, a, b)
    
    # MOSTRAR RESULTADOS
    print("RESULTADOS ESTADÍSTICOS")
    print()
    
    print("ESTADÍSTICAS DESCRIPTIVAS:")
    print(f"Número de observaciones (n):        {n}")
    print(f"Media de X (x̄):                     {mean_X:.6f}")
    print(f"Media de Y (ȳ):                     {mean_Y:.6f}")
    print()
    
    print(" DESVIACIÓN ESTÁNDAR TOTAL (s):")

    print(f"Fórmula: s = √[Σ(yi - ȳ)² / (n-1)]")
    print(f"Resultado: {desv_std:.6f}")
    print()
    
    print("ERROR ESTÁNDAR DEL ESTIMADO (Sₑ):")
 
    print(f"Fórmula: Sₑ = √[Σ(yi - ŷi)² / (n-2)]")
    print(f"Resultado: {see:.6f}")
    print()
    
    print(" COEFICIENTE DE CORRELACIÓN (r):")
 
    print(f"Fórmula: r = Σ(xi-x̄)(yi-ȳ) / √[Σ(xi-x̄)²·Σ(yi-ȳ)²]")
    print(f"Resultado: {r:.6f}")
    if r > 0.7:
        print(f"Interpretación: Correlación positiva FUERTE")
    elif r > 0.3:
        print(f"Interpretación: Correlación positiva MODERADA")
    elif r > -0.3:
        print(f"Interpretación: Correlación DÉBIL")
    elif r > -0.7:
        print(f"Interpretación: Correlación negativa MODERADA")
    else:
        print(f"Interpretación: Correlación negativa FUERTE")
    print()
    
    print("COEFICIENTE DE DETERMINACIÓN (R²):")
 
    print(f"Fórmula: R² = r²")
    print(f"Resultado: {r2:.6f}")
    print(f"Porcentaje: {r2*100:.2f}% de la varianza es explicada")
    print()
    
    print("ECUACIÓN DE REGRESIÓN LINEAL:")
 
    print(f"y = {a:.6f} + {b:.6f}x")
    print(f"Intercepto (a): {a:.6f}")
    print(f"Pendiente (b):  {b:.6f}")
    print()
    
 
    print("CÁLCULO COMPLETADO")
 

# EJECUTAR EL PROGRAMA
if __name__ == "__main__":
    # Solicitar la ubicación completa del archivo
    print("CALCULADORA DE ESTADÍSTICAS")
 

    
    archivo = input("Ubicación del archivo: ").strip()
    
    # Limpiar comillas si el usuario las incluyó
    if archivo.startswith('"') and archivo.endswith('"'):
        archivo = archivo[1:-1]
    if archivo.startswith("'") and archivo.endswith("'"):
        archivo = archivo[1:-1]
    
    print()
    
    try:
        calcular_estadisticas(archivo)
    except FileNotFoundError:
        print(f"Error: No se encontró el archivo en la ubicación especificada")
        print(f"Ruta buscada: {archivo}")
        